"""
export_excel.py — Academic Excel Workbook Generator.

This module compiles the backtest results into a publication-ready Excel
workbook. It strictly focuses on empirical results, omitting zero-trade 
anomalies and highlighting parameter combinations that meet the >=50% 
accuracy benchmark.
"""

import logging
from pathlib import Path
from typing import Dict, Any, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

# ── Style Utilities ────────────────────────────────────────────────────────


def _fill(hex_code: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_code)


def _border() -> Border:
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)


def _hdr(ws: Any, row: int, fill: PatternFill) -> None:
    for cell in ws[row]:
        if cell.value is not None:
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = fill
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True)
            cell.border = _border()


def _style_row(ws: Any, row: int, even: bool) -> None:
    bg = "F5F7FA" if even else "FFFFFF"
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill = _fill(bg)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _border()
            cell.font = Font(size=9)


def _autowidth(ws: Any, mn: int = 9, mx: int = 30) -> None:
    for col in ws.columns:
        ltr = get_column_letter(col[0].column)
        w = max((len(str(c.value or "")) for c in col), default=mn)
        ws.column_dimensions[ltr].width = min(max(w + 2, mn), mx)


def _freeze(ws: Any, row: int = 2) -> None:
    ws.freeze_panes = ws.cell(row=row, column=1)


def _color_cell(cell: Any, value: float, invert: bool = False) -> None:
    if isinstance(value, (int, float)):
        is_good = value > 0 if not invert else value < 0
        if is_good:
            cell.fill = _fill("C8E6C9")
            cell.font = Font(color="1B5E20", size=9, bold=True)
        elif value < 0:
            cell.fill = _fill("FFCDD2")
            cell.font = Font(color="B71C1C", size=9, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _border()


def _color_winrate(cell: Any, wr: float) -> None:
    """Colors the win rate cell green (>=50%), yellow (40-50%), or red (<40%)."""
    if wr >= 50:
        cell.fill = _fill("1B5E20")
        cell.font = Font(color="FFFFFF", size=9, bold=True)
    elif wr >= 40:
        cell.fill = _fill("FFF9C4")
        cell.font = Font(color="6D4C00", size=9, bold=True)
    else:
        cell.fill = _fill("FFCDD2")
        cell.font = Font(color="B71C1C", size=9, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _border()


# ── Column Definitions ───────────────────────────────────────────────────
SUMMARY_COLS: List[str] = [
    "Market", "Timeframe", "EMA Pair", "Fast EMA", "Slow EMA",
    "Angle Threshold", "Total Trades", "Long Trades", "Short Trades",
    "Winning", "Losing", "Win Rate (%)", "Total Profit ($)", "Avg Profit ($)",
    "Profit Factor", "Sharpe Ratio", "Max Drawdown (%)", "Final Equity ($)"
]

TRADE_COLS: List[str] = [
    "#", "Direction", "Entry Date", "Exit Date",
    "Entry Price", "Exit Price", "Qty",
    "PnL ($)", "Return (%)", "Result"
]


class ExcelExporter:
    """Generates the publication-ready Excel workbook containing experimental results."""

    COLORS: Dict[str, str] = {
        "summary": "1F2937",
        "fifty":   "10B981",
        "trades":  "6B7280"
    }

    def export(
        self,
        results: Dict[str, Dict[str, Dict[Any, Dict[str, Any]]]],
        initial_equity: float,
        output_path: Path,
    ) -> None:
        wb = Workbook()
        wb.remove(wb.active)

        self._summary(wb, results)
        self._fifty_plus(wb, results)
        self._trade_logs(wb, results)

        wb.save(output_path)
        log.info(f"Excel Export Complete: {output_path.name}")

    def _summary(self, wb: Workbook, results: dict) -> None:
        """Aggregates all active combinations (Total Trades > 0)."""
        ws = wb.create_sheet("Empirical Summary")
        fill = _fill(self.COLORS["summary"])

        active = sum(1 for tf_dict in results.values() for pair_dict in tf_dict.values()
                     for m in pair_dict.values() if m.get("total_trades", 0) > 0)

        ws.append(
            [f"EMA Angle Optimization — Full Empirical Dataset ({active} valid combos)"])
        ws["A1"].font = Font(bold=True, size=13, color=self.COLORS["summary"])
        ws.merge_cells(f"A1:{get_column_letter(len(SUMMARY_COLS))}1")

        ws.append(SUMMARY_COLS)
        _hdr(ws, 2, fill)

        WR_COL = SUMMARY_COLS.index("Win Rate (%)") + 1
        PR_COL = SUMMARY_COLS.index("Total Profit ($)") + 1
        DD_COL = SUMMARY_COLS.index("Max Drawdown (%)") + 1

        ri = 3
        for asset, tf_dict in results.items():
            for tf, pair_dict in tf_dict.items():
                sorted_items = sorted(pair_dict.items(), key=lambda kv: kv[1].get(
                    "total_profit", -1e18), reverse=True)
                for (f, s, thr), m in sorted_items:
                    if m.get("total_trades", 0) == 0:
                        continue

                    wr = round(m.get("win_rate", 0.0), 2)
                    ws.append([
                        asset, tf, f"EMA({f},{s})", f, s, thr,
                        m["total_trades"], m["long_trades"], m["short_trades"],
                        m["winning_trades"], m["losing_trades"],
                        wr,
                        round(m["total_profit"], 2), round(m["avg_profit"], 2),
                        round(m["profit_factor"], 2), round(
                            m["sharpe_ratio"], 4),
                        round(m["max_drawdown"], 2), round(
                            m["final_equity"], 2),
                    ])
                    _style_row(ws, ri, ri % 2 == 0)
                    _color_winrate(ws.cell(ri, WR_COL), wr)
                    _color_cell(ws.cell(ri, PR_COL), m["total_profit"])
                    _color_cell(ws.cell(ri, DD_COL),
                                m["max_drawdown"], invert=True)
                    ri += 1

        _autowidth(ws)
        _freeze(ws)
        ws.sheet_properties.tabColor = self.COLORS["summary"]

    def _fifty_plus(self, wb: Workbook, results: dict) -> None:
        """Filters the summary to only include statistically significant, high-accuracy combos."""
        ws = wb.create_sheet("High Accuracy (Win Rate >= 50%)")
        fill = _fill(self.COLORS["fifty"])

        fifty_count = sum(1 for tf_dict in results.values() for pair_dict in tf_dict.values()
                          for m in pair_dict.values() if m.get("total_trades", 0) > 0 and m.get("win_rate", 0) >= 50)

        ws.append(
            [f"Optimized Parameters (Win Rate >= 50%): {fifty_count} configurations found"])
        ws["A1"].font = Font(bold=True, size=13, color=self.COLORS["fifty"])
        ws.merge_cells(f"A1:{get_column_letter(len(SUMMARY_COLS))}1")

        ws.append(SUMMARY_COLS)
        _hdr(ws, 2, fill)

        WR_COL = SUMMARY_COLS.index("Win Rate (%)") + 1
        PR_COL = SUMMARY_COLS.index("Total Profit ($)") + 1
        DD_COL = SUMMARY_COLS.index("Max Drawdown (%)") + 1

        ri = 3
        for asset, tf_dict in results.items():
            for tf, pair_dict in tf_dict.items():
                sorted_items = sorted(pair_dict.items(), key=lambda kv: kv[1].get(
                    "total_profit", -1e18), reverse=True)
                for (f, s, thr), m in sorted_items:
                    wr = round(m.get("win_rate", 0.0), 2)
                    if m.get("total_trades", 0) == 0 or wr < 50:
                        continue

                    ws.append([
                        asset, tf, f"EMA({f},{s})", f, s, thr,
                        m["total_trades"], m["long_trades"], m["short_trades"],
                        m["winning_trades"], m["losing_trades"],
                        wr,
                        round(m["total_profit"], 2), round(m["avg_profit"], 2),
                        round(m["profit_factor"], 2), round(
                            m["sharpe_ratio"], 4),
                        round(m["max_drawdown"], 2), round(
                            m["final_equity"], 2),
                    ])
                    _style_row(ws, ri, ri % 2 == 0)
                    _color_winrate(ws.cell(ri, WR_COL), wr)
                    _color_cell(ws.cell(ri, PR_COL), m["total_profit"])
                    _color_cell(ws.cell(ri, DD_COL),
                                m["max_drawdown"], invert=True)
                    ri += 1

        _autowidth(ws)
        _freeze(ws)
        ws.sheet_properties.tabColor = self.COLORS["fifty"]

    def _trade_logs(self, wb: Workbook, results: dict) -> None:
        """Outputs the individual transaction ledger for the most profitable parameter set per asset/timeframe."""
        fill = _fill(self.COLORS["trades"])

        for asset, tf_dict in results.items():
            for tf, pair_dict in tf_dict.items():
                if not pair_dict:
                    continue

                best_key = max(pair_dict.keys(), key=lambda k: pair_dict[k].get(
                    "total_profit", -1e18))
                m = pair_dict[best_key]
                if m.get("total_trades", 0) == 0:
                    continue

                f, s, thr = best_key
                ws = wb.create_sheet(f"{asset[:3]}_{tf}_Logs")

                title = f"Transaction Ledger | {asset} ({tf}) | EMA({f},{s}) | Angle >= {thr}° | Profit: ${m['total_profit']:,.2f}"
                ws.append([title])
                ws["A1"].font = Font(bold=True, size=11,
                                     color=self.COLORS["trades"])
                ws.merge_cells(f"A1:{get_column_letter(len(TRADE_COLS))}1")

                ws.append(TRADE_COLS)
                _hdr(ws, 2, fill)

                PNL_COL = TRADE_COLS.index("PnL ($)") + 1
                RET_COL = TRADE_COLS.index("Return (%)") + 1
                RES_COL = TRADE_COLS.index("Result") + 1

                for i, t in enumerate(m["trades"], 1):
                    pnl = t["pnl"]
                    res = "WIN" if pnl > 0 else "LOSS"
                    dt_fmt = "%Y-%m-%d %H:%M"

                    ws.append([
                        i, t["direction"],
                        t["entry_date"].strftime(dt_fmt) if pd.notnull(
                            t["entry_date"]) else "",
                        t["exit_date"].strftime(dt_fmt) if pd.notnull(
                            t["exit_date"]) else "",
                        t["entry_price"], t["exit_price"], t["qty"],
                        round(pnl, 2), round(t["return_pct"], 2), res
                    ])

                    ri = i + 2
                    _style_row(ws, ri, ri % 2 == 0)
                    _color_cell(ws.cell(ri, PNL_COL), pnl)
                    _color_cell(ws.cell(ri, RET_COL), t["return_pct"])

                    res_cell = ws.cell(ri, RES_COL)
                    res_cell.font = Font(
                        color="1B5E20" if res == "WIN" else "B71C1C", bold=True)

                _autowidth(ws)
                _freeze(ws)
                ws.sheet_properties.tabColor = "9CA3AF"
